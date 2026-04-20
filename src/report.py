"""
Excel report generator.

Sheets produced
---------------

1. "Top Candidates"        — best CSP across all tickers, ranked by yield.
2. "<TICKER>"              — one sheet per watchlist entry, all surviving puts.
3. "T-Bill Matching"       — bucket -> yield mapping used in the run.
4. "Settings"              — snapshot of the settings.yaml that was active.
5. "Watchlist"             — snapshot of the watchlist that was scanned.
"""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .option_selector import CSPCandidate
from .tbill import TBillMatch, TBillMatcher
from .watchlist import Settings, WatchlistEntry


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
NUMERIC_ALIGN = Alignment(horizontal="right")
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# Column definitions: (header, attr, number_format, width)
TOP_COLUMNS = [
    ("Symbol",         "symbol",            "@",          10),
    ("Expiry",         "expiry_fmt",        "@",          12),
    ("DTE",            "dte",               "0",           7),
    ("Strike",         "strike",            "#,##0.00",   10),
    ("Spot",           "underlying_price",  "#,##0.00",   10),
    ("Moneyness",      "moneyness",         "0.00%",      11),
    ("Bid",            "bid",               "0.00",        8),
    ("Ask",            "ask",               "0.00",        8),
    ("Mid",            "mid",               "0.00",        8),
    ("Spread%",        "spread_pct",        "0.00%",      10),
    ("IV",             "iv",                "0.00%",       8),
    ("Delta",          "delta",             "0.0000",     10),
    ("Open Int.",      "open_interest",     "#,##0",      10),
    ("Volume",         "volume",            "#,##0",      10),
    ("Cash Req. (USD)", "cash_required",    "#,##0",      14),
    ("Premium (USD)",  "premium",           "#,##0.00",   13),
    ("Ann. Yield",     "annualized_yield",  "0.00%",      11),
    ("Breakeven",      "breakeven",         "#,##0.00",   11),
    ("T-Bill Bucket",  "tbill_bucket",      "0",          13),
    ("T-Bill Yield",   "tbill_yield",       "0.00%",      12),
    ("T-Bill Interest","tbill_interest",    "#,##0.00",   14),
    ("Total P/L if expires worthless", "total_if_otm", "#,##0.00", 22),
    ("Total Yield (Prem+TBill)",       "total_yield",  "0.00%",    20),
]


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def write_report(
    output_path: Path,
    candidates_by_ticker: dict[str, list[CSPCandidate]],
    tbill_matches: dict[str, dict[str, TBillMatch]],   # ticker -> expiry -> match
    matcher: TBillMatcher,
    watchlist: list[WatchlistEntry],
    settings: Settings,
    run_ts: datetime | None = None,
) -> Path:
    run_ts = run_ts or datetime.utcnow()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Default sheet -> repurpose as Top Candidates
    ws_top = wb.active
    ws_top.title = "Top Candidates"

    # ---- Aggregate top across tickers --------------------------------------
    aggregated: list[tuple[str, CSPCandidate, TBillMatch]] = []
    for ticker, candidates in candidates_by_ticker.items():
        for c in candidates:
            tb = tbill_matches.get(ticker, {}).get(c.expiry)
            aggregated.append((ticker, c, tb))
    aggregated.sort(key=lambda x: x[1].annualized_yield, reverse=True)

    _write_candidate_sheet(
        ws=ws_top,
        title=f"CSP Scanner — Top Candidates  ({run_ts.strftime('%Y-%m-%d %H:%M UTC')})",
        rows=aggregated,
    )

    # ---- Per-ticker detail sheets ------------------------------------------
    for ticker, candidates in candidates_by_ticker.items():
        ws = wb.create_sheet(title=_safe_sheet_name(ticker))
        rows = [
            (ticker, c, tbill_matches.get(ticker, {}).get(c.expiry))
            for c in candidates
        ]
        _write_candidate_sheet(
            ws=ws,
            title=f"{ticker} — Cash-Secured Puts",
            rows=rows,
        )

    # ---- T-Bill matching sheet --------------------------------------------
    _write_tbill_sheet(wb.create_sheet("T-Bill Matching"), matcher, tbill_matches)

    # ---- Settings & Watchlist snapshots -----------------------------------
    _write_settings_sheet(wb.create_sheet("Settings"), settings, run_ts)
    _write_watchlist_sheet(wb.create_sheet("Watchlist"), watchlist)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _write_candidate_sheet(
    ws,
    title: str,
    rows: list[tuple[str, CSPCandidate, TBillMatch | None]],
) -> None:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TOP_COLUMNS))

    header_row = 3
    for ci, (header, _, _, width) in enumerate(TOP_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    if not rows:
        ws.cell(row=header_row + 1, column=1, value="(no candidates passed the filters)")
        return

    for ri, (ticker, cand, tbill) in enumerate(rows, start=header_row + 1):
        view = _candidate_view(cand, tbill)
        for ci, (_, attr, fmt, _) in enumerate(TOP_COLUMNS, start=1):
            value = view.get(attr)
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.number_format = fmt
            cell.border = CELL_BORDER
            if isinstance(value, (int, float)):
                cell.alignment = NUMERIC_ALIGN

    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)


def _candidate_view(c: CSPCandidate, tbill: TBillMatch | None) -> dict:
    view = asdict(c)
    view["expiry_fmt"] = _fmt_expiry(c.expiry)
    if tbill:
        view["tbill_bucket"] = tbill.bucket_days
        view["tbill_yield"] = tbill.yield_pct
        interest = tbill.interest_on(c.cash_required)
        view["tbill_interest"] = interest
        # P/L if put expires worthless: keep the premium + earn T-Bill interest
        view["total_if_otm"] = c.premium + interest
        # Combined yield on cash-secured capital, annualized
        view["total_yield"] = c.annualized_yield + tbill.yield_pct
    else:
        view["tbill_bucket"] = None
        view["tbill_yield"] = None
        view["tbill_interest"] = None
        view["total_if_otm"] = c.premium
        view["total_yield"] = c.annualized_yield
    return view


def _write_tbill_sheet(ws, matcher: TBillMatcher, tbill_matches: dict) -> None:
    ws.cell(row=1, column=1, value="T-Bill Matching").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ["Bucket (days)", "Yield (decimal)", "Source", "Note"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60

    for ri, (bucket, (yld, source)) in enumerate(sorted(matcher._yield_cache.items()), start=4):
        ws.cell(row=ri, column=1, value=bucket)
        c_yld = ws.cell(row=ri, column=2, value=yld)
        c_yld.number_format = "0.00%"
        ws.cell(row=ri, column=3, value=source)
        note = "Live IB quote" if source == "live" else "Fallback yield from settings.yaml"
        ws.cell(row=ri, column=4, value=note)

    # Detail of which expiry mapped to which bucket
    detail_start = ws.max_row + 3
    ws.cell(row=detail_start, column=1, value="Per-Expiry Detail").font = TITLE_FONT
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=4)
    sub_headers = ["Ticker", "Expiry", "DTE", "Bucket"]
    for ci, h in enumerate(sub_headers, start=1):
        cell = ws.cell(row=detail_start + 2, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    r = detail_start + 3
    for ticker, by_expiry in tbill_matches.items():
        for expiry, m in sorted(by_expiry.items()):
            ws.cell(row=r, column=1, value=ticker)
            ws.cell(row=r, column=2, value=_fmt_expiry(expiry))
            ws.cell(row=r, column=3, value=m.dte)
            ws.cell(row=r, column=4, value=m.bucket_days)
            r += 1


def _write_settings_sheet(ws, settings: Settings, run_ts: datetime) -> None:
    ws.cell(row=1, column=1, value="Settings Snapshot").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Run: {run_ts.strftime('%Y-%m-%d %H:%M UTC')}")

    rows: list[tuple[str, object]] = []
    rows.append(("[ib]", ""))
    for k, v in asdict(settings.ib).items():
        rows.append((f"  {k}", v))
    rows.append(("[options]", ""))
    for k, v in asdict(settings.options).items():
        rows.append((f"  {k}", v))
    rows.append(("[tbill]", ""))
    for k, v in asdict(settings.tbill).items():
        rows.append((f"  {k}", v))
    rows.append(("[report]", ""))
    for k, v in asdict(settings.report).items():
        rows.append((f"  {k}", v))

    for ri, (k, v) in enumerate(rows, start=4):
        ws.cell(row=ri, column=1, value=k)
        ws.cell(row=ri, column=2, value=str(v))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _write_watchlist_sheet(ws, entries: list[WatchlistEntry]) -> None:
    ws.cell(row=1, column=1, value="Watchlist").font = TITLE_FONT
    headers = ["Symbol", "Exchange", "Currency", "Max Strike", "Max Contracts", "Notes"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    widths = [10, 10, 10, 14, 16, 50]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, e in enumerate(entries, start=4):
        ws.cell(row=ri, column=1, value=e.symbol)
        ws.cell(row=ri, column=2, value=e.exchange)
        ws.cell(row=ri, column=3, value=e.currency)
        ws.cell(row=ri, column=4, value=e.max_strike).number_format = "#,##0.00"
        ws.cell(row=ri, column=5, value=e.max_contracts)
        ws.cell(row=ri, column=6, value=e.notes)


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------


def _fmt_expiry(yyyymmdd: str) -> str:
    if not yyyymmdd or len(yyyymmdd) < 8:
        return yyyymmdd
    return f"{yyyymmdd[:4]}-{yyyymmdd[4:6]}-{yyyymmdd[6:8]}"


def _safe_sheet_name(name: str) -> str:
    # Excel limits: 31 chars, no [ ] : * ? / \
    bad = '[]:*?/\\'
    cleaned = "".join("_" if ch in bad else ch for ch in name)
    return cleaned[:31] or "sheet"
